import csv
import io
import ipaddress
import re
import socket
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook


TARGET_FIELDS = ("linkedin", "facebook", "instagram", "phone", "email")


def validate_public_url(url: str) -> str:
    parsed = urlparse(url if "://" in url else f"https://{url}")
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("A valid public HTTP(S) website is required")
    for item in socket.getaddrinfo(parsed.hostname, parsed.port or 443):
        address = ipaddress.ip_address(item[4][0])
        if not address.is_global:
            raise ValueError("Private or unsafe network destinations are not allowed")
    return parsed.geturl()


def crawl_company(url: str) -> tuple[dict, str]:
    current = validate_public_url(url)
    for _ in range(4):
        response = requests.get(current, timeout=15, allow_redirects=False, headers={"User-Agent": "AutoLead/1.0"})
        if response.is_redirect:
            current = validate_public_url(response.headers["location"])
            continue
        response.raise_for_status()
        soup = BeautifulSoup(response.text[:2_000_000], "html.parser")
        links = [a.get("href", "") for a in soup.find_all("a", href=True)]
        result = {field: None for field in TARGET_FIELDS}
        for link in links:
            lower = link.lower()
            if "linkedin.com" in lower and not result["linkedin"]: result["linkedin"] = link
            if "facebook.com" in lower and not result["facebook"]: result["facebook"] = link
            if "instagram.com" in lower and not result["instagram"]: result["instagram"] = link
            if lower.startswith("mailto:") and not result["email"]: result["email"] = link[7:].split("?")[0]
            if lower.startswith("tel:") and not result["phone"]: result["phone"] = link[4:]
        if not result["email"]:
            match = re.search(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", soup.get_text(" "))
            result["email"] = match.group(0) if match else None
        return result, current
    raise ValueError("Too many redirects")


def parse_upload(filename: str, content: bytes) -> list[tuple[str, list[str], list[dict]]]:
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        return [("Sheet1", headers, [dict(row) for row in reader])]
    if filename.lower().endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=True)
        result = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                result.append((sheet.title, [], [])); continue
            headers = [str(value or f"Column {index + 1}") for index, value in enumerate(rows[0])]
            values = [{headers[i]: row[i] if i < len(row) else None for i in range(len(headers))} for row in rows[1:]]
            result.append((sheet.title, headers, values))
        return result
    raise ValueError("Only CSV and XLSX files are supported")


def safe_cell(value):
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def workbook_bytes(sheets) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    used = set()
    for name, headers, rows in sheets:
        clean = re.sub(r"[:\\/?*\[\]]", "_", name)[:31] or "Sheet"
        base, counter = clean, 2
        while clean in used:
            clean = f"{base[:27]} {counter}"; counter += 1
        used.add(clean)
        sheet = workbook.create_sheet(clean)
        sheet.append(headers)
        for row in rows:
            sheet.append([safe_cell(row.get(header)) for header in headers])
    output = io.BytesIO(); workbook.save(output)
    return output.getvalue()
